import json
import openpyxl

from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from datetime import datetime, timedelta
from django.core.exceptions import ValidationError
from django.db import models
from openpyxl.styles import Alignment
from django.http import HttpResponse
from robots.models import Robot

@csrf_exempt
def create_robot(request):
    if request.method != "POST":
        return JsonResponse({"error": "Only POST requetst are allowed"}, status=405)
    
    try:
        data = json.loads(request.body)

        required_fields = ["model", "version", "created"]
        for field in required_fields:
            if field not in data:
                return JsonResponse({"error": f"Missing required field: {field}"}, status=400)
        
        try:
            created_datetime = datetime.strptime(data["created"], "%Y-%m-%d %H:%M:%S")
        except ValueError:
            return JsonResponse({"error": "Invalid datetime format for 'created'. Expected format: YYYY-MM-DD HH:MM:SS."}, status=400)

        robot = Robot(
            model=data["model"],
            version=data["version"],
            created=created_datetime
        )

        robot.full_clean()
        robot.save()
        return JsonResponse({"message": "Robot created successfully.", "robot_id": robot.id}, status=201)
    
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON format."}, status=400)
    except ValidationError as e:
        return JsonResponse({"error": "Validation error.", "details": e.message_dict}, status=400)
    except Exception as e:
        return JsonResponse({"error": "An unexpected error occurred.", "details": str(e)}, status=500)

def generate_production_report():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Temporary"

    end_date = datetime.now()
    start_date = end_date - timedelta(days=7)

    robots = (
        Robot.objects.filter(created__range=(start_date, end_date))
        .values("model", "version")
        .annotate(total=models.Count("id"))
        .order_by("model", "version")
    )

    grouped_data = {}
    for robot in robots:
        model = robot["model"]
        if model not in grouped_data:
            grouped_data[model] = []
        grouped_data[model].append({"version": robot["version"], "total": robot["total"]})
    
    for model, versions in grouped_data.items():
        new_sheet = workbook.create_sheet(title=model)
        new_sheet.append(["Модель", "Версия", "Количество за неделю"])
        for col in ["A", "B", "C"]:
            new_sheet[f"{col}1"].alignment = Alignment(horizontal="center", vertical="center")
        
        for version_data in versions:
            new_sheet.append([model, version_data["version"], version_data["total"]])
    
    if 'Temporary' in workbook.sheetnames:
        del workbook['Temporary']
    
    return workbook

def download_production_report(request):
    workbook = generate_production_report()
    
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="production_report.xlsx"'
    
    workbook.save(response)
    return response